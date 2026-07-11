"""
Excel Service (CRM)

This service manages all low-level CRUD operations on the Excel sheet (data/Leads.xlsx) 
which acts as our CRM (Customer Relationship Management) database.

It uses 'openpyxl' under the hood to open, read, edit, and save the Excel workbook.
It is completely dynamic, loading headers on initialization to map column titles to 1-indexed column numbers.

Core Responsibilities:
---------------------
1. Load workbook: Opens data/Leads.xlsx safely using case-insensitive path lookup.
2. Load headers: Dynamically reads the first row to match column names (e.g., 'Lead ID') to column indices.
3. Read pending leads: Returns a list of dicts representing leads matching Status='New' or Call Status='Pending'.
4. Find leads: Iterates through the rows to find a matching 'Lead ID' or 'Last Call SID'.
5. Update leads: Updates cells dynamically in the workbook for a specific row and saves the changes.

Workbook Header Column Mapping:
-------------------------------
- Lead ID (int): Primary key.
- Name (str): Lead contact name.
- Phone (str): Cleaned and formatted dynamically on dial.
- Status (str): Lifecycle status (New / Pending / Contacted / Closed).
- Call Status (str): Call lifecycle (Pending / Ringing / Connected / Completed / No Answer / Busy / Failed).
- Last Call SID (str): Twilio Call identifier (used to update status via webhooks).
- Conversation Summary (str): AI-generated call report.
- Objections Raised (str): Customer objections captured by AI.
"""

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from app.core.constants import (
    LEAD_ID,
    STATUS,
    LAST_CONTACTED,
    RETRY_COUNT,
    LAST_CALL_SID,
    CALL_STATUS,
)


class ExcelService:
    """
    Service to manage read, find, and update operations on the Excel CRM spreadsheet.
    """

    def __init__(self):
        """
        Initializes the ExcelService by loading the workbook and mapping the column headers.
        """
        # Find the Excel file (case-insensitive for cross-platform compatibility)
        from glob import glob
        excel_files = glob("data/[Ll][Ee][Aa][Dd][Ss].xlsx")
        if excel_files:
            self.file_path = Path(excel_files[0])
        else:
            self.file_path = Path("data/Leads.xlsx")

        print(f"\nLoading Excel File : {self.file_path.resolve()}")

        # openpyxl workbook loader
        self.workbook = load_workbook(self.file_path)

        # Active sheet (usually the first sheet)
        self.sheet = self.workbook.active

        # Load headers to dynamically map column titles (e.g. 'Phone') to indices (e.g. 4)
        self.headers = self._load_headers()

    # --------------------------------------------------
    # Internal Helpers
    # --------------------------------------------------

    def _load_headers(self):
        """
        Read the header row and build a
        dictionary of column names.
        """

        headers = {}

        for column in range(1, self.sheet.max_column + 1):

            value = self.sheet.cell(
                row=1,
                column=column,
            ).value

            headers[str(value).strip()] = column

        return headers

    def save(self):
        """
        Save workbook.
        """

        self.workbook.save(self.file_path)

    def close(self):
        """
        Close workbook.
        """

        self.workbook.close()

    # --------------------------------------------------
    # Read Operations
    # --------------------------------------------------

    def get_lead_by_row(self, row):

        lead = {}

        for header, column in self.headers.items():

            lead[header] = self.sheet.cell(
                row=row,
                column=column,
            ).value

        return lead

    def get_pending_leads(self):
        """
        Return leads that are ready to be called.

        A lead is "pending" if:
        - Call Status is 'Pending'  (scheduled but not yet dialled)
        - OR Status is 'New'        (brand-new lead, never contacted)
        """

        leads = []

        call_status_column = self.headers.get(CALL_STATUS)
        status_column = self.headers.get(STATUS)

        for row in range(2, self.sheet.max_row + 1):

            lead_id_val = self.sheet.cell(
                row=row,
                column=self.headers[LEAD_ID],
            ).value

            if lead_id_val is None:
                continue

            call_status = ""
            if call_status_column:
                val = self.sheet.cell(row=row, column=call_status_column).value
                call_status = str(val).strip().lower() if val else ""

            status = ""
            if status_column:
                val = self.sheet.cell(row=row, column=status_column).value
                status = str(val).strip().lower() if val else ""

            # A lead is pending calling if call_status is explicitly 'pending', 
            # or if it is brand-new (status 'new') and has never been called (call_status is empty/none/nan/null).
            is_pending_call_status = call_status == "pending"
            is_new_and_never_called = status == "new" and call_status in ["", "none", "nan", "null"]

            if is_pending_call_status or is_new_and_never_called:
                leads.append(self.get_lead_by_row(row))

        return leads

    # --------------------------------------------------
    # Find Lead
    # --------------------------------------------------

    def find_row_by_lead_id(
        self,
        lead_id,
    ):
        """
        Find Excel row using Lead ID.

        Handles int, float and string IDs.
        """

        lead_column = self.headers[LEAD_ID]

        print("\nSearching Lead...")

        print(f"Requested Lead ID : {lead_id}")

        for row in range(2, self.sheet.max_row + 1):

            value = self.sheet.cell(
                row=row,
                column=lead_column,
            ).value

            print(
                f"Row {row} -> {value}"
            )

            if value is None:
                continue

            if str(value).strip() == str(lead_id).strip():

                print(f"Lead Found at Excel Row {row}")

                return row

        print("Lead Not Found")

        return None

    # --------------------------------------------------
    # Read/Find All
    # --------------------------------------------------

    def get_all_leads(self):
        """
        Return all leads.
        """
        leads = []
        for row in range(2, self.sheet.max_row + 1):
            lead_id_val = self.sheet.cell(row=row, column=self.headers[LEAD_ID]).value
            if lead_id_val is not None:
                leads.append(self.get_lead_by_row(row))
        return leads

    def get_lead_by_id(self, lead_id):
        """
        Return one lead by ID.
        """
        row = self.find_row_by_lead_id(lead_id)
        if row is None:
            return None
        return self.get_lead_by_row(row)

    def find_row_by_call_sid(
        self,
        call_sid,
    ):
        """
        Find Excel row using Last Call SID.
        """
        if LAST_CALL_SID not in self.headers:
            return None

        call_sid_column = self.headers[LAST_CALL_SID]

        for row in range(2, self.sheet.max_row + 1):
            value = self.sheet.cell(
                row=row,
                column=call_sid_column,
            ).value

            if value is None:
                continue

            if str(value).strip() == str(call_sid).strip():
                return row

        return None

    # --------------------------------------------------
    # Update Lead
    # --------------------------------------------------

    def update_lead_by_row(self, row, updates):
        """
        Update a lead row directly.
        """
        print(f"\nUpdating Excel Row {row}...\n")

        for field, value in updates.items():
            if field not in self.headers:
                print(f"Skipping unknown column : {field}")
                continue

            column = self.headers[field]
            self.sheet.cell(row=row, column=column).value = value
            print(f"{field} -> {value}")

        from datetime import timezone, timedelta
        ist = timezone(timedelta(hours=5, minutes=30))
        self.sheet.cell(
            row=row,
            column=self.headers[LAST_CONTACTED],
        ).value = datetime.now(ist).strftime("%Y-%m-%d %H:%M:%S")

        self.save()
        print("\nWorkbook Saved Successfully")

    def update_lead(
        self,
        lead_id,
        updates,
    ):
        """
        Update one lead using Lead ID.
        """
        row = self.find_row_by_lead_id(lead_id)
        if row is None:
            raise ValueError(f"Lead {lead_id} not found.")

        self.update_lead_by_row(row, updates)

    def update_lead_by_call_sid(
        self,
        call_sid,
        updates,
    ):
        """
        Update one lead using Last Call SID.
        """
        row = self.find_row_by_call_sid(call_sid)
        if row is None:
            raise ValueError(f"Lead with Last Call SID {call_sid} not found.")

        self.update_lead_by_row(row, updates)
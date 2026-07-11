# -*- coding: utf-8 -*-
"""
Import all 4 n8n workflows via REST API and apply Excel Data Validation on startup.
Runs inside the FastAPI Docker container.
"""
import urllib.request
import urllib.error
import json
import http.cookiejar
import os
import time
from pathlib import Path

# Detect if running inside Docker or on host
if os.path.exists("/.dockerenv"):
    N8N_URL = "http://n8n:5678"
else:
    N8N_URL = "http://localhost:5678"

EMAIL = "ashishsharma12512@gmail.com"
PASSWORD = "0187As231012"

WORKFLOWS = [
    "n8n/lead_scheduler.json",
    "n8n/retry_calls.json",
    "n8n/meeting_reminder.json",
    "n8n/daily_summary.json",
]

def wait_for_n8n(url, timeout=120):
    print(f"[SETUP] Waiting for n8n to start at {url}...")
    start_time = time.time()
    while time.time() - start_time < timeout:
        try:
            req = urllib.request.Request(url, method="GET")
            urllib.request.urlopen(req, timeout=2)
            print("[SETUP] [OK] n8n is online!")
            return True
        except urllib.error.HTTPError as e:
            print(f"[SETUP] [OK] n8n is online (responded with HTTP {e.code})!")
            return True
        except Exception:
            time.sleep(2)
    print("[SETUP] [WARN] n8n did not start within the timeout period. Trying to connect anyway...")
    return False

def normalize_workflow(wf_data):
    nodes = []
    for node in wf_data.get("nodes", []):
        node_copy = node.copy()
        node_copy.pop("position", None)
        node_copy.pop("id", None)
        nodes.append(node_copy)
    nodes = sorted(nodes, key=lambda x: x.get("name", ""))
    connections = wf_data.get("connections", {})
    return {"nodes": nodes, "connections": connections}

def is_different(local_data, remote_data):
    local_norm = normalize_workflow(local_data)
    remote_norm = normalize_workflow(remote_data)
    l_str = json.dumps(local_norm, sort_keys=True)
    r_str = json.dumps(remote_norm, sort_keys=True)
    return l_str != r_str

def activate_workflow(opener, cookie_header, wf_id, version_id):
    if not version_id:
        return False
    act_req = urllib.request.Request(
        f"{N8N_URL}/rest/workflows/{wf_id}/activate",
        data=json.dumps({"versionId": version_id}).encode(),
        headers={
            "Content-Type": "application/json",
            "Cookie": cookie_header,
        },
        method="POST",
    )
    try:
        opener.open(act_req)
        print(f"[SETUP] Activated workflow {wf_id}")
        return True
    except Exception as ae:
        print(f"[SETUP] Activate failed for {wf_id}: {ae}")
        return False

def deactivate_workflow(opener, cookie_header, wf_id):
    deact_req = urllib.request.Request(
        f"{N8N_URL}/rest/workflows/{wf_id}/deactivate",
        data=b"",
        headers={
            "Content-Type": "application/json",
            "Cookie": cookie_header,
        },
        method="POST",
    )
    try:
        opener.open(deact_req)
        print(f"[SETUP] Deactivated workflow {wf_id}")
        return True
    except Exception as de:
        print(f"[SETUP] Deactivate failed for {wf_id}: {de}")
        return False

def apply_excel_validation():
    import openpyxl
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    file_path = "/app/data/Leads.xlsx"
    if not os.path.exists(file_path):
        file_path = "data/Leads.xlsx"
        
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        headers = {}
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val:
                headers[str(val).strip()] = col
                
        validations = {
            "Status": {
                "list": "New,Pending,Contacted,Closed,Interested,Not Interested,Pricing Inquiry,Warm Lead,Engaged",
                "prompt": "Please select a lifecycle status",
                "error": "Please select a valid lifecycle status from the list"
            },
            "Call Status": {
                "list": "Pending,Queued,Ringing,Connected,Completed,No Answer,Busy,Failed,Canceled",
                "prompt": "Please select a call lifecycle status",
                "error": "Please select a valid call lifecycle status from the list"
            },
            "Lead Qualification": {
                "list": "Qualified,Warm,Cold,Not Qualified,Need More Information",
                "prompt": "Please select a qualification level",
                "error": "Please select a valid qualification level from the list"
            }
        }
        
        ws.data_validations.dataValidation = []
        
        for col_name, config in validations.items():
            if col_name in headers:
                col_idx = headers[col_name]
                col_letter = get_column_letter(col_idx)
                cell_range = f"{col_letter}2:{col_letter}1000"
                
                formula = f'"{config["list"]}"'
                dv = DataValidation(
                    type="list",
                    formula1=formula,
                    allow_blank=True
                )
                
                dv.promptTitle = f"{col_name} Selection"
                dv.prompt = config["prompt"]
                dv.errorTitle = "Invalid Selection"
                dv.error = config["error"]
                
                ws.add_data_validation(dv)
                dv.add(cell_range)
                
        wb.save(file_path)
        wb.close()
        print("[SETUP] [OK] Excel data validation applied to Leads.xlsx columns G, H, and K.")
    except Exception as e:
        print(f"[SETUP] [WARNING] Could not apply Excel data validation: {e}")

def run_setup():
    # 1. Apply Excel Data Validation
    apply_excel_validation()

    # 2. Sync Workflows to n8n
    if not wait_for_n8n(N8N_URL):
        return

    cj = http.cookiejar.CookieJar()
    opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(cj))

    # Login
    login_payload = json.dumps({"emailOrLdapLoginId": EMAIL, "password": PASSWORD}).encode()
    login_req = urllib.request.Request(
        f"{N8N_URL}/rest/login",
        data=login_payload,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    try:
        resp = opener.open(login_req)
        user = json.loads(resp.read().decode())
        print("[SETUP] Logged in to n8n as: " + user['data']['email'])
    except Exception as e:
        print(f"[SETUP] [ERR] n8n login failed: {e}")
        return

    cookie_header = "; ".join(f"{c.name}={c.value}" for c in cj)

    # Fetch and clean duplicates
    existing_workflows = {}
    duplicates_to_delete = []
    list_req = urllib.request.Request(
        f"{N8N_URL}/rest/workflows",
        headers={"Cookie": cookie_header},
        method="GET"
    )
    try:
        r = opener.open(list_req)
        list_data = json.loads(r.read().decode())
        
        grouped = {}
        for wf in list_data.get("data", []):
            name = wf["name"]
            grouped.setdefault(name, []).append(wf)
            
        for name, wfs in grouped.items():
            non_archived = [w for w in wfs if not w.get("isArchived")]
            if non_archived:
                keep_wf = non_archived[0]
                existing_workflows[name] = keep_wf
                for w in wfs:
                    if w["id"] != keep_wf["id"]:
                        duplicates_to_delete.append((name, w["id"], w.get("isArchived", False)))
            else:
                for w in wfs:
                    duplicates_to_delete.append((name, w["id"], True))
        
        for name, dup_id, is_archived in duplicates_to_delete:
            if not is_archived:
                archive_req = urllib.request.Request(
                    f"{N8N_URL}/rest/workflows/{dup_id}/archive",
                    data=b"",
                    headers={"Cookie": cookie_header},
                    method="POST",
                )
                try:
                    opener.open(archive_req)
                except:
                    pass
                    
            del_req = urllib.request.Request(
                f"{N8N_URL}/rest/workflows/{dup_id}",
                headers={"Cookie": cookie_header},
                method="DELETE"
            )
            try:
                opener.open(del_req)
                print(f"[SETUP] Deleted duplicate workflow: {name}")
            except:
                pass
    except Exception as e:
        print(f"[SETUP] [WARNING] Failed to clean duplicates: {e}")

    # Import / Update Workflows
    for wf_path in WORKFLOWS:
        # Check relative or absolute paths inside container
        full_path = Path("/app") / wf_path
        if not full_path.exists():
            full_path = Path(wf_path)
            
        if not full_path.exists():
            print(f"[SETUP] [WARN] File not found: {wf_path}")
            continue

        try:
            with open(full_path, "r", encoding="utf-8") as f:
                wf_data = json.load(f)

            wf_name = wf_data.get("name", wf_path)
            existing_wf = existing_workflows.get(wf_name)

            if existing_wf:
                existing_id = existing_wf["id"]
                active_status = existing_wf.get("active", False)
                
                detail_req = urllib.request.Request(
                    f"{N8N_URL}/rest/workflows/{existing_id}",
                    headers={"Cookie": cookie_header},
                    method="GET"
                )
                try:
                    detail_resp = opener.open(detail_req)
                    detailed_remote_wf = json.loads(detail_resp.read().decode()).get("data", {})
                except Exception as e:
                    detailed_remote_wf = existing_wf

                version_id = detailed_remote_wf.get("versionId")

                if not is_different(wf_data, detailed_remote_wf):
                    print(f"[SETUP] Workflow up to date: {wf_name}")
                    if not active_status:
                        activate_workflow(opener, cookie_header, existing_id, version_id)
                    continue

                print(f"[SETUP] Updating workflow: {wf_name}")
                wf_data.pop("id", None)
                payload = json.dumps(wf_data).encode()
                req = urllib.request.Request(
                    f"{N8N_URL}/rest/workflows/{existing_id}",
                    data=payload,
                    headers={
                        "Content-Type": "application/json",
                        "Cookie": cookie_header,
                    },
                    method="PATCH",
                )
                action_name = "Updated"
                wf_id = existing_id
            else:
                wf_data.pop("id", None)
                payload = json.dumps(wf_data).encode()
                req = urllib.request.Request(
                    f"{N8N_URL}/rest/workflows",
                    data=payload,
                    headers={
                        "Content-Type": "application/json",
                        "Cookie": cookie_header,
                    },
                    method="POST",
                )
                action_name = "Created"
                wf_id = None

            r = opener.open(req)
            result = json.loads(r.read().decode())
            if not wf_id:
                wf_id = result.get("data", {}).get("id", "?")
            new_version_id = result.get("data", {}).get("versionId")
            print(f"[SETUP] [OK] {action_name}: {wf_name} (id={wf_id})")

            deactivate_workflow(opener, cookie_header, wf_id)
            time.sleep(1)
            activate_workflow(opener, cookie_header, wf_id, new_version_id)

        except Exception as e:
            print(f"[SETUP] [ERR] Import/update failed for {wf_path}: {e}")

    print("[SETUP] Setup complete. n8n workflows and CRM sheet initialized.")
